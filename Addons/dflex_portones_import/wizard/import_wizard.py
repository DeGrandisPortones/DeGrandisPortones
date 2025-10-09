from odoo import api, fields, models
from odoo.exceptions import UserError
import base64, io, re, unicodedata, csv

HEADER_NV_CANDIDATES = [
    'nv', 'n v', 'numero de venta', 'n° de venta', 'nº de venta',
    'nro venta', 'nro de venta', 'número de venta', 'n° venta', 'venta n°', 'nota de venta'
]

def _norm(s):
    return (str(s) if s is not None else '').strip()

def _lower(s):
    return _norm(s).lower()

def _slug(s):
    s = _norm(s)
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'[^a-zA-Z0-9]+', '_', s).strip('_').lower()
    if not s:
        s = 'col'
    if s and s[0].isdigit():
        s = 'c_' + s
    return s[:50]

def _label(h1, h2):
    h1n, h2n = _norm(h1), _norm(h2)
    if h1n and h2n:
        return (h1n[:3].capitalize() + '. ' + h2n).strip()
    return (h1n or h2n)

def _is_xlsx(data):
    return data.startswith(b'PK\x03\x04')

def _is_xls(data):
    return data.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1')

def _ffill(values):
    out, last = [], ''
    for v in values:
        if _norm(v):
            last = _norm(v)
            out.append(last)
        else:
            out.append(last)
    return out

class DflexPortonImportWizard(models.TransientModel):
    _name = 'dflex.porton.import.wizard'
    _description = 'Importar portones desde Excel/CSV (PRINCIPAL, fila1+fila2 encabezados, datos desde fila3)'

    file = fields.Binary(required=True, string='Archivo XLS/XLSX/CSV')
    filename = fields.Char(string='Nombre de archivo')

    def _read_excel(self, content, filename):
        data = base64.b64decode(content or b'')
        if not data:
            raise UserError('El archivo está vacío.')
        sheet_name = 'PRINCIPAL'
        fname = (filename or '').lower()

        # --- CSV directo ---
        if fname.endswith('.csv'):
            text = data.decode('utf-8', errors='replace')
            rows = list(csv.reader(io.StringIO(text)))
            if len(rows) < 3:
                raise UserError('CSV inválido: se esperan 2 filas de encabezado y datos desde la 3.')
            headers1 = [ _norm(x) for x in rows[0] ]
            headers2 = [ _norm(x) for x in rows[1] ]
            headers1 = _ffill(headers1)
            data_rows = rows[2:]
            return headers1, headers2, data_rows

        # --- XLSX ---
        if _is_xlsx(data) or fname.endswith('.xlsx'):
            try:
                import openpyxl
            except Exception as e:
                raise UserError('Falta dependencia openpyxl para .xlsx: %s' % e)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            wanted = None
            for sn in wb.sheetnames:
                if _lower(sn) == _lower(sheet_name):
                    wanted = sn
                    break
            if not wanted:
                raise UserError('No se encontró la hoja "%s".' % sheet_name)
            ws = wb[wanted]
            max_col = ws.max_column or 0
            headers1 = [ _norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1, max_col=max_col)) ]
            headers2 = [ _norm(c.value) for c in next(ws.iter_rows(min_row=2, max_row=2, max_col=max_col)) ]
            headers1 = _ffill(headers1)
            rows = []
            for r in ws.iter_rows(min_row=3, max_col=max_col, values_only=True):
                rows.append([ v for v in r ])
            return headers1, headers2, rows

        # --- XLS ---
        if _is_xls(data) or fname.endswith('.xls'):
            try:
                import xlrd
            except Exception:
                raise UserError('Tu archivo es .XLS y esta instancia no trae xlrd. Abrí el archivo y guardalo como .XLSX o subí CSV.')
            book = xlrd.open_workbook(file_contents=data)
            wanted = None
            for sn in book.sheet_names():
                if _lower(sn) == _lower(sheet_name):
                    wanted = sn
                    break
            if not wanted:
                raise UserError('No se encontró la hoja "%s".' % sheet_name)
            sh = book.sheet_by_name(wanted)
            ncols = sh.ncols
            headers1 = [ _norm(sh.cell_value(0, c)) for c in range(ncols) ]
            headers2 = [ _norm(sh.cell_value(1, c)) for c in range(ncols) ]
            headers1 = _ffill(headers1)
            rows = []
            for r in range(2, sh.nrows):
                rows.append([ sh.cell_value(r, c) for c in range(ncols) ])
            return headers1, headers2, rows

        raise UserError('Formato no reconocido. Subí XLSX/XLS o CSV.')

    def _build_columns(self, headers1, headers2):
        cols = []
        n = max(len(headers1), len(headers2))
        for idx in range(n):
            h1 = headers1[idx] if idx < len(headers1) else ''
            h2 = headers2[idx] if idx < len(headers2) else ''
            if re.match(r'^\s*columna\b', _lower(h1)) or re.match(r'^\s*columna\b', _lower(h2)):
                continue
            if not _norm(h1) and not _norm(h2):
                continue
            lbl = _label(h1, h2) or ('Columna %s' % (idx+1))
            tech = ('x_%s_%s' % (_slug(h1), _slug(h2))).strip('_')
            if not tech.startswith('x_'):
                tech = 'x_' + tech
            cols.append({'index': idx, 'label': lbl, 'tech': tech[:63], 'h1': h1, 'h2': h2})
        seen = {}
        for c in cols:
            base = c['tech']
            k = base
            i = 1
            while k in seen:
                i += 1
                k = (base[:60] + '_' + str(i))[:63]
            c['tech'] = k
            seen[k] = True
        return cols

    def _ensure_fields(self, columns):
        Model = self.env['ir.model'].sudo()
        Fields = self.env['ir.model.fields'].sudo()
        model = Model.search([('model', '=', 'dflex.porton')], limit=1)
        if not model:
            raise UserError('No se encontró el modelo dflex.porton.')
        out = []
        for col in columns:
            name = col['tech']
            field = Fields.search([('model_id', '=', model.id), ('name', '=', name)], limit=1)
            if not field:
                field = Fields.create({
                    'name': name,
                    'model_id': model.id,
                    'ttype': 'char',
                    'field_description': col['label'],
                    'store': True,
                })
            out.append({'name': name, 'label': col['label']})
        return out

    def _upsert_dynamic_form_view(self, fields_meta):
        View = self.env['ir.ui.view'].sudo()
        lines = ['                  <field name="%s" string="%s"/>' % (f['name'], f['label']) for f in fields_meta]
        field_xml = "\n".join(lines) or '                  <separator string="(sin columnas detectadas)"/>'
        arch = """
        <form string="Portón">
          <sheet>
            <group>
              <field name="name"/>
              <field name="import_id"/>
              <field name="source_row"/>
            </group>
            <notebook>
              <page string="Datos importados">
                <group col="4">
{}
                </group>
              </page>
              <page string="JSON">
                <field name="specs_json" widget="json"/>
              </page>
            </notebook>
          </sheet>
        </form>
        """.format(field_xml)
        rec = View.search([('model', '=', 'dflex.porton'), ('name', '=', 'dflex.porton.form.auto')], limit=1)
        vals = {'arch_db': arch, 'type': 'form', 'priority': 1}
        if rec:
            rec.write(vals)
        else:
            View.create({'name': 'dflex.porton.form.auto', 'model': 'dflex.porton', **vals})

    def _find_nv_column(self, columns):
        for c in columns:
            text = ' '.join([_lower(c['label']), _lower(c['h1']), _lower(c['h2'])])
            for cand in HEADER_NV_CANDIDATES:
                if cand in text:
                    return c
        return None

    def action_import(self):
        self.ensure_one()
        h1, h2, data_rows = self._read_excel(self.file, self.filename or 'import.xlsx')
        columns = self._build_columns(h1, h2)
        if not columns:
            raise UserError('No se detectaron columnas válidas en PRINCIPAL (revisá fila 1 y 2).')
        meta = self._ensure_fields(columns)
        idx2tech = { c['index']: c['tech'] for c in columns }
        nv_col = self._find_nv_column(columns)

        batch = self.env['dflex.porton.import'].create({
            'name': self.filename or 'Importación',
            'file_name': self.filename,
            'total_rows': len(data_rows),
            'note': 'Hoja/Archivo: PRINCIPAL | Columnas: %s' % ', '.join([c['label'] for c in columns]),
            'state': 'draft',
        })

        Porton = self.env['dflex.porton']
        Spec = self.env['dflex.porton.spec']
        created_ids = []
        for r_idx, row in enumerate(data_rows, start=3):
            values, specs = {}, {}
            # Soporta row como lista de strings (CSV) o lista de celdas (XLSX/XLS)
            for i, cell in enumerate(row):
                if i in idx2tech:
                    v = '' if cell is None else str(cell)
                    values[idx2tech[i]] = v
                    specs[idx2tech[i]] = v
            name_val = None
            if nv_col and nv_col['index'] < len(row):
                nv_raw = row[nv_col['index']]
                name_val = _norm(nv_raw)
            if not name_val:
                name_val = 'Sin NV (fila %s)' % r_idx
            rec = Porton.create({
                'name': name_val,
                'import_id': batch.id,
                'source_row': r_idx,
                'specs_json': specs,
                **values,
            })
            if specs:
                Spec.create([{'porton_id': rec.id, 'key': k, 'value': v} for k, v in specs.items()])
            created_ids.append(rec.id)

        batch.state = 'done'
        self._upsert_dynamic_form_view([{'name': f['name'], 'label': f['label']} for f in meta])
        action = self.env.ref('dflex_portones_import.action_dflex_porton').read()[0]
        action['domain'] = [('id', 'in', created_ids)]
        return action