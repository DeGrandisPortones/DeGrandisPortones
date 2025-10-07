# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
from io import BytesIO

ALLOWED_COMPANY_NAME = "De Grandis Portones"

try:
    from openpyxl import load_workbook
except Exception as e:
    load_workbook = None

class GateImportWizard(models.TransientModel):
    _name = "x.gate.import.wizard"
    _description = "Importar Portones desde Excel"

    file_data = fields.Binary("Archivo Excel (.xlsx)", required=True)
    file_name = fields.Char("Nombre de archivo")
    sheet_name = fields.Char("Nombre de hoja", default="PRINCIPAL")
    header_row = fields.Integer("Fila de encabezados", default=1)
    data_start_row = fields.Integer("Primera fila de datos", default=2)
    create_batch = fields.Boolean("Crear lote de importación", default=True)

    def _check_company(self):
        if self.env.company.name != ALLOWED_COMPANY_NAME:
            raise UserError(_("Esta función está habilitada solo para la empresa '%s'. Empresa actual: '%s'")
                            % (ALLOWED_COMPANY_NAME, self.env.company.name))

    def action_import(self):
        self._check_company()

        if not load_workbook:
            raise UserError(_("Falta instalar 'openpyxl'. Ejecute: pip install openpyxl"))

        if not self.file_data:
            raise UserError(_("Debe subir un archivo .xlsx."))

        data = base64.b64decode(self.file_data)
        wb = load_workbook(filename=BytesIO(data), data_only=True)
        sheet_name = self.sheet_name or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise UserError(_("La hoja '%s' no existe. Hojas disponibles: %s") % (sheet_name, ', '.join(wb.sheetnames)))
        ws = wb[sheet_name]

        hdr_row = self.header_row or 1
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[hdr_row]]

        batch = None
        if self.create_batch:
            batch = self.env['x.gate.import.batch'].create({
                'name': self.file_name or 'Importación',
                'file_data': self.file_data,
                'file_name': self.file_name,
                'sheet_name': sheet_name,
                'company_id': self.env.company.id,
            })

        start = self.data_start_row or (hdr_row + 1)
        created = 0
        Gate = self.env['x.gate.spec']

        for row in ws.iter_rows(min_row=start, values_only=True):
            if all(v in (None, '', ' ') for v in row):
                continue

            row_dict = {}
            for idx, val in enumerate(row):
                col = headers[idx] if idx < len(headers) else ''
                sval = '' if val is None else str(val).strip()
                row_dict[col] = sval

            name = row_dict.get('Nota de Venta') or row_dict.get('NOTA DE VENTA') or ''
            cliente = row_dict.get('CLIENTE +3000') or row_dict.get('CLIENTE') or ''
            if name and cliente:
                rec_name = f"{name} - {cliente}"
            elif name:
                rec_name = str(name)
            elif cliente:
                rec_name = str(cliente)
            else:
                rec_name = _('Portón importado')

            Gate.create({
                'name': rec_name,
                'import_batch_id': batch.id if batch else False,
                'data_json': row_dict,
                'company_id': self.env.company.id,
            })
            created += 1

        if batch:
            batch.row_count = created

        msg = _('Se importaron %s filas desde la hoja %s.') % (created, sheet_name)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {'title': _('Importación completada'), 'message': msg, 'type': 'success', 'sticky': False},
        }
